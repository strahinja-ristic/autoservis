"""
AutoServis — Uvoz podataka iz Excel-a u bazu
Pokretanje: python uvoz_excel.py
Potrebno: pip install openpyxl
"""

import sqlite3
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Nedostaje biblioteka. Pokreni: pip install openpyxl")
    sys.exit(1)


APPDATA = os.environ.get("APPDATA", os.path.expanduser("~"))
DEFAULT_DB = os.path.join(APPDATA, "AutoServis", "autoservis.db")

TABELE = ["klijenti", "vozila", "artikli", "sabloni_usluga"]

# Kolone koje se preskacu automatski (ID, arhiviran itd.)
PRESKOK = {"id", "arhiviran", "arhivirano"}

# Marke sa dva reci (za automatski split)
DVORECHE_MARKE = {
    "ALFA ROMEO", "LAND ROVER", "ASTON MARTIN", "ROLLS ROYCE",
    "MERCEDES BENZ", "GREAT WALL",
}

# Opis kolona za lakse mapiranje
OPIS_KOLONA = {
    # klijenti
    "tip": "tip klijenta — vrednost mora biti 'Fizicko' ili 'Pravno'",
    "ime": "ime (za fizicka lica)",
    "prezime": "prezime (za fizicka lica)",
    "naziv_firme": "naziv firme (za pravna lica)",
    "pib": "PIB (poreski identifikacioni broj)",
    "maticni_broj": "maticni broj firme",
    "adresa": "adresa",
    "telefon": "telefon",
    "email": "email adresa",
    "napomena": "napomena",
    # vozila
    "klijent_id": "ID klijenta iz baze (broj) — mora vec postojati u klijenti tabeli",
    "marka": "marka vozila (npr. Volkswagen)",
    "model": "model vozila (npr. Golf 5)",
    "godiste": "godiste vozila (broj, npr. 2008)",
    "registracija": "registarska oznaka",
    "broj_sasije": "broj sasije (VIN)",
    "kilometraza": "kilometraza (broj)",
    # artikli
    "naziv": "naziv artikla/usluge",
    "jedinica_mere": "jedinica mere (kom, l, kg...)",
    "kolicina": "kolicina na stanju (broj)",
    "nabavna_cena": "nabavna cena bez PDV (broj)",
    "prodajna_cena": "prodajna cena bez PDV (broj)",
    "minimalna_kolicina": "minimalna kolicina (broj)",
    # sabloni_usluga
    "cena": "cena usluge (broj)",
}


def unos(tekst, podrazumevano=None):
    if podrazumevano:
        odgovor = input(f"{tekst} [{podrazumevano}]: ").strip()
        return odgovor if odgovor else podrazumevano
    return input(f"{tekst}: ").strip()


def get_kolone_tabele(conn, tabela):
    rows = conn.execute(f"PRAGMA table_info({tabela})").fetchall()
    # (cid, name, type, notnull, dflt_value, pk)
    return rows


def izaberi_kolonu(poruka, excel_kolone, dozvoli_kombinovano=False):
    """Prikaz numerisane liste Excel kolona, korisnik bira broj ili Enter za preskok.
    Ako dozvoli_kombinovano=True, nudi i opciju 'K' za kolonu koja sadrzi i marka i model."""
    print(f"\n  {poruka}")
    for i, kol in enumerate(excel_kolone, 1):
        print(f"    {i:>2}. {kol}")
    print(f"     0. [preskok]")
    if dozvoli_kombinovano:
        print(f"     K. [ova kolona sadrzi MARKU i MODEL zajedno]")

    while True:
        izbor = input("  Izbor: ").strip()
        if izbor == "" or izbor == "0":
            return None
        if dozvoli_kombinovano and izbor.upper() == "K":
            return "__KOMBINOVANO__"
        if izbor.isdigit():
            idx = int(izbor) - 1
            if 0 <= idx < len(excel_kolone):
                return excel_kolone[idx]
        if izbor in excel_kolone:
            return izbor
        hint = f"broj od 1 do {len(excel_kolone)}, 0 za preskok"
        if dozvoli_kombinovano:
            hint += ", ili K za kombinovano"
        print(f"  Nevazeci unos. Unesi {hint}.")


def split_marka_model(vrednost, reci_marke=1):
    """Splituje 'SKODA SCALA' na ('SKODA', 'SCALA'). reci_marke=2 za 'ALFA ROMEO GIULIA'."""
    if not vrednost:
        return None, None
    delovi = str(vrednost).strip().split()
    if len(delovi) <= reci_marke:
        return " ".join(delovi), None
    marka = " ".join(delovi[:reci_marke])
    model = " ".join(delovi[reci_marke:])
    return marka, model


def main():
    print("=" * 55)
    print("  AutoServis — Uvoz podataka iz Excel-a")
    print("=" * 55)

    # --- 1. Putanja do baze ---
    print(f"\nPodrazumevana baza: {DEFAULT_DB}")
    db_path = unos("Pritisni Enter za podrazumevano, ili unesi drugu putanju", DEFAULT_DB)
    if not os.path.exists(db_path):
        print(f"\nGreska: Baza nije pronadjena na: {db_path}")
        sys.exit(1)

    conn = sqlite3.connect(db_path)
    print("  OK — baza ucitana.")

    # --- 2. Excel fajl ---
    print()
    excel_path = unos("Putanja do Excel fajla (mozes prevuci fajl ovde)").strip('"').strip("'")
    if not os.path.exists(excel_path):
        print(f"\nGreska: Fajl nije pronadjen: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path)
    sheets = wb.sheetnames

    if len(sheets) == 1:
        sheet_name = sheets[0]
        print(f"  Koristim sheet: {sheet_name}")
    else:
        print(f"\nPronasao sam {len(sheets)} sheet-ova:")
        for i, s in enumerate(sheets, 1):
            print(f"  {i}. {s}")
        while True:
            izbor = unos("Koji sheet koristiti? (broj ili naziv)")
            if izbor.isdigit() and 1 <= int(izbor) <= len(sheets):
                sheet_name = sheets[int(izbor) - 1]
                break
            elif izbor in sheets:
                sheet_name = izbor
                break
            print("  Nevazeci izbor.")

    ws = wb[sheet_name]

    # Citaj zaglavlje (prvi red)
    excel_kolone = []
    for cell in ws[1]:
        if cell.value is not None:
            excel_kolone.append(str(cell.value).strip())

    if not excel_kolone:
        print("\nGreska: Prvi red Excel-a je prazan — ocekujem zaglavlje (nazive kolona) u prvom redu.")
        sys.exit(1)

    print(f"\n  Pronadjene kolone u Excel-u ({len(excel_kolone)}):")
    for i, k in enumerate(excel_kolone, 1):
        print(f"    {i:>2}. {k}")

    # --- 3. Tabela ---
    print(f"\nDostupne tabele:")
    for i, t in enumerate(TABELE, 1):
        print(f"  {i}. {t}")
    while True:
        izbor = unos("U koju tabelu uvoziš?")
        if izbor.isdigit() and 1 <= int(izbor) <= len(TABELE):
            tabela = TABELE[int(izbor) - 1]
            break
        elif izbor in TABELE:
            tabela = izbor
            break
        print("  Nevazeci izbor.")

    print(f"\n  Izabrana tabela: {tabela}")

    # --- 4. Mapiranje kolona ---
    db_kolone = get_kolone_tabele(conn, tabela)

    print(f"\n{'=' * 55}")
    print(f"  Mapiranje kolona — tabela: {tabela}")
    print(f"  Za svaku kolonu baze izaberi odgovarajucu Excel kolonu.")
    print(f"  Unesi broj ili 0/Enter za preskok.")
    print(f"{'=' * 55}")

    mapiranje = {}        # db_kolona -> excel_kolona
    split_kolona = None   # naziv Excel kolone koja sadrzi marka+model zajedno
    split_reci = 1        # koliko reci je marka

    preskoci_model = False  # ako je marka+model kombinovano, preskoči pitanje za model

    for col in db_kolone:
        col_name = col[1]
        col_type = col[2]
        is_pk = col[5]

        if is_pk or col_name in PRESKOK:
            continue

        if col_name == "model" and preskoci_model:
            print(f"\n  model -> [izvuceno automatski iz iste kolone kao marka]  ✓")
            continue

        opis = OPIS_KOLONA.get(col_name, "")
        poruka = f"Kolona baze: {col_name} ({col_type})"
        if opis:
            poruka += f"\n  Opis: {opis}"

        je_marka = (col_name == "marka" and tabela == "vozila")
        excel_kol = izaberi_kolonu(poruka, excel_kolone, dozvoli_kombinovano=je_marka)

        if excel_kol == "__KOMBINOVANO__":
            # Pitaj korisnika za Excel kolonu i broj reci marke
            print(f"\n  Koja Excel kolona sadrzi marka+model zajedno?")
            excel_kol_izvor = izaberi_kolonu("Izaberi kolonu:", excel_kolone)
            if not excel_kol_izvor:
                continue

            while True:
                reci_str = input("  Koliko reci je marka? (1 = SKODA/VW/BMW, 2 = ALFA ROMEO): ").strip()
                if reci_str in ("1", "2"):
                    split_reci = int(reci_str)
                    break
                print("  Unesi 1 ili 2.")

            split_kolona = excel_kol_izvor
            mapiranje["marka"] = excel_kol_izvor
            mapiranje["model"] = excel_kol_izvor
            preskoci_model = True
            print(f"  -> {excel_kol_izvor} (split na {split_reci} {'rec' if split_reci == 1 else 'reci'})  ✓")

        elif excel_kol:
            mapiranje[col_name] = excel_kol
            print(f"  -> {excel_kol}  ✓")

    if not mapiranje:
        print("\nNema mapiranih kolona. Izlazim.")
        sys.exit(0)

    # --- 5. Potvrda ---
    print(f"\n{'=' * 55}")
    print(f"  Mapiranje koje ce biti primenjeno:")
    prikazano = set()
    for db_kol, ex_kol in mapiranje.items():
        if split_kolona and db_kol in ("marka", "model"):
            kljuc = f"split:{ex_kol}"
            if kljuc not in prikazano:
                print(f"    Excel [{ex_kol}]  ->  baza [marka] + [model]  (split na {split_reci} {'rec' if split_reci == 1 else 'reci'})")
                prikazano.add(kljuc)
        else:
            print(f"    Excel [{ex_kol}]  ->  baza [{db_kol}]")

    ukupno_redova = ws.max_row - 1
    print(f"\n  Tabela: {tabela}")
    print(f"  Redova za uvoz: ~{ukupno_redova}")
    potvrda = unos("\nNastavljam sa uvozom? (da/ne)")
    if potvrda.lower() not in ("da", "d", "yes", "y"):
        print("Otkazano.")
        sys.exit(0)

    # --- 6. Uvoz ---
    header_index = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1]) if cell.value is not None}

    db_kols = list(mapiranje.keys())
    # Ukloni duplikat 'model' ako je kombinovano (oba marka i model se čitaju iz split_kolona)
    db_kols_unique = list(dict.fromkeys(db_kols))
    ex_kols = [mapiranje[k] for k in db_kols_unique]
    placeholders = ", ".join(["?"] * len(db_kols_unique))
    col_names_sql = ", ".join(db_kols_unique)
    insert_sql = f"INSERT INTO {tabela} ({col_names_sql}) VALUES ({placeholders})"

    ubaceno = 0
    preskoceno = 0
    greske = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        vrednosti = []
        for db_kol, ex_kol in zip(db_kols_unique, ex_kols):
            col_idx = header_index.get(ex_kol)
            val = row[col_idx].value if col_idx is not None else None
            if isinstance(val, str):
                val = val.strip() if val.strip() != "" else None

            # Split marka+model
            if split_kolona and ex_kol == split_kolona and db_kol == "marka":
                marka, _ = split_marka_model(val, split_reci)
                vrednosti.append(marka)
            elif split_kolona and ex_kol == split_kolona and db_kol == "model":
                _, model = split_marka_model(val, split_reci)
                vrednosti.append(model)
            else:
                vrednosti.append(val)

        # Preskoči potpuno prazne redove
        if all(v is None for v in vrednosti):
            preskoceno += 1
            continue

        try:
            conn.execute(insert_sql, vrednosti)
            ubaceno += 1
        except sqlite3.Error as e:
            print(f"  Red {row_idx}: Greska — {e}")
            greske += 1

    conn.commit()
    conn.close()

    print(f"\n{'=' * 55}")
    print(f"  Gotovo!")
    print(f"  Ubaceno redova : {ubaceno}")
    print(f"  Praznih redova : {preskoceno}")
    print(f"  Gresaka        : {greske}")
    print(f"{'=' * 55}")
    if greske > 0:
        print(f"\n  Napomena: Redovi sa greskama nisu ubaceni.")
        print(f"  Najcesci uzrok: obavezno polje je prazno ili klijent_id ne postoji.")
    input("\nPritisni Enter za izlaz...")


if __name__ == "__main__":
    main()
