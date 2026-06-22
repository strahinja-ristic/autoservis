"""
AutoServis — Uvoz artikala iz Excel-a
Pokretanje: python uvoz_artikala.py
Potrebno: pip install openpyxl

Ocekivani format Excel-a (zaglavlje u prvom redu):
  SIFRA | NAZIV | JM | CENA

  SIFRA  -> ignorise se (autoincrement u bazi)
  NAZIV  -> naziv artikla
  JM     -> jedinica mere
  CENA   -> prodajna cena (nabavna se postavlja na 0)
"""

import sqlite3
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Nedostaje biblioteka. Pokreni: pip install openpyxl")
    sys.exit(1)


APPDATA = os.environ.get("APPDATA", os.path.expanduser("~"))
DEFAULT_DB = os.path.join(APPDATA, "AutoServis", "autoservis.db")


def unos(tekst, podrazumevano=None):
    if podrazumevano:
        odg = input(f"{tekst} [{podrazumevano}]: ").strip()
        return odg if odg else podrazumevano
    return input(f"{tekst}: ").strip()


def main():
    print("=" * 55)
    print("  AutoServis — Uvoz artikala")
    print("=" * 55)

    # --- Baza ---
    print(f"\nPodrazumevana baza: {DEFAULT_DB}")
    db_path = unos("Pritisni Enter za podrazumevano, ili unesi drugu putanju", DEFAULT_DB)
    if not os.path.exists(db_path):
        print(f"\nGreska: Baza nije pronadjena: {db_path}")
        sys.exit(1)

    conn = sqlite3.connect(db_path)
    print("  OK — baza ucitana.")

    # --- Excel ---
    print()
    excel_path = unos("Putanja do Excel fajla").strip('"').strip("'")
    if not os.path.exists(excel_path):
        print(f"\nGreska: Fajl nije pronadjen: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path)
    sheets = wb.sheetnames

    if len(sheets) == 1:
        sheet_name = sheets[0]
        print(f"  Koristim sheet: {sheet_name}")
    else:
        print(f"\nSheet-ovi: {', '.join(sheets)}")
        sheet_name = unos("Koji sheet?")

    ws = wb[sheet_name]

    # Procitaj zaglavlje
    zaglavlje = {str(c.value).strip().upper(): i for i, c in enumerate(ws[1]) if c.value}
    print(f"\n  Pronadjene kolone: {', '.join(zaglavlje.keys())}")

    def nadji(* kandidati):
        for k in kandidati:
            if k.upper() in zaglavlje:
                return zaglavlje[k.upper()]
        return None

    idx_naziv = nadji("NAZIV", "IME", "ARTIKAL")
    idx_jm    = nadji("JM", "JEDINICA MERE", "JEDINICA", "MJ")
    idx_cena  = nadji("CENA", "CIJENA", "PRODAJNA CENA", "PRODAJNA")

    if idx_naziv is None:
        print("\nGreska: Kolona NAZIV nije pronadjena.")
        sys.exit(1)

    print(f"\n  Mapiranje:")
    print(f"    NAZIV -> naziv")
    print(f"    JM    -> jedinica_mere  {'(kolona ' + str(idx_jm + 1) + ')' if idx_jm is not None else '(nije nadjena, bice prazno)'}")
    print(f"    CENA  -> prodajna_cena  {'(kolona ' + str(idx_cena + 1) + ')' if idx_cena is not None else '(nije nadjena, bice 0)'}")
    print(f"    nabavna_cena -> 0  (fiksno)")

    ukupno = ws.max_row - 1
    print(f"\n  Redova za uvoz: ~{ukupno}")
    potvrda = unos("Nastavljam? (da/ne)")
    if potvrda.lower() not in ("da", "d", "yes", "y"):
        print("Otkazano.")
        sys.exit(0)

    def cel(row, idx):
        if idx is None:
            return None
        val = row[idx].value
        if isinstance(val, str):
            val = val.strip() or None
        return val

    ubaceno = 0
    preskoceno = 0
    greske = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        naziv = cel(row, idx_naziv)

        if not naziv:
            preskoceno += 1
            continue

        jm   = cel(row, idx_jm)
        cena = cel(row, idx_cena)

        try:
            cena_f = float(str(cena).replace(",", ".")) if cena is not None else 0.0
        except ValueError:
            print(f"  Red {row_idx}: Nevazeca cena '{cena}', postavljam na 0.")
            cena_f = 0.0

        try:
            conn.execute(
                """INSERT INTO artikli (naziv, jedinica_mere, prodajna_cena, nabavna_cena)
                   VALUES (?, ?, ?, 0)""",
                (naziv, jm, cena_f)
            )
            ubaceno += 1
        except sqlite3.Error as e:
            print(f"  Red {row_idx}: Greska — {e}")
            greske += 1

    conn.commit()
    conn.close()

    print(f"\n{'=' * 55}")
    print(f"  Gotovo!")
    print(f"  Ubaceno artikala : {ubaceno}")
    print(f"  Praznih redova   : {preskoceno}")
    print(f"  Gresaka          : {greske}")
    print(f"{'=' * 55}")
    input("\nPritisni Enter za izlaz...")


if __name__ == "__main__":
    main()
