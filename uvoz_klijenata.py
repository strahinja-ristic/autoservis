"""
AutoServis — Uvoz / azuriranje klijenata (pravna lica) iz Excel-a
Pokretanje: python uvoz_klijenata.py
Potrebno: pip install openpyxl

Ocekivani format Excel-a (zaglavlje u prvom redu):
  FIRMA | ADRESA | GRAD | PIB | MATICNI

  - Ako klijent sa tim nazivom vec postoji -> azurira ga (adresa, grad, PIB, maticni)
  - Ako ne postoji -> dodaje novog (tip=Pravno)
  - ADRESA i GRAD se spajaju u jedno polje: "Adresa, Grad"
"""

import sqlite3
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Nedostaje biblioteka. Pokreni: pip install openpyxl")
    sys.exit(1)


APPDATA = os.environ.get("APPDATA", os.path.expanduser("~"))
DEFAULT_DB = os.path.join(APPDATA, "AutoServis", "autoservis.db")


def unos(tekst, podrazumevano=None):
    if podrazumevano:
        odg = input(f"{tekst} [{podrazumevano}]: ").strip()
        return odg if odg else podrazumevano
    return input(f"{tekst}: ").strip()


def main():
    print("=" * 55)
    print("  AutoServis — Uvoz klijenata (pravna lica)")
    print("=" * 55)

    # --- Baza ---
    print(f"\nPodrazumevana baza: {DEFAULT_DB}")
    db_path = unos("Pritisni Enter za podrazumevano, ili unesi drugu putanju", DEFAULT_DB)
    if not os.path.exists(db_path):
        print(f"\nGreska: Baza nije pronadjena: {db_path}")
        sys.exit(1)

    conn = sqlite3.connect(db_path)
    print("  OK — baza ucitana.")

    # --- Excel ---
    print()
    excel_path = unos("Putanja do Excel fajla").strip('"').strip("'")
    if not os.path.exists(excel_path):
        print(f"\nGreska: Fajl nije pronadjen: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path)
    sheets = wb.sheetnames

    if len(sheets) == 1:
        sheet_name = sheets[0]
        print(f"  Koristim sheet: {sheet_name}")
    else:
        print(f"\nSheet-ovi: {', '.join(sheets)}")
        sheet_name = unos("Koji sheet?")

    ws = wb[sheet_name]

    zaglavlje = {str(c.value).strip().upper(): i for i, c in enumerate(ws[1]) if c.value}
    print(f"\n  Pronadjene kolone: {', '.join(zaglavlje.keys())}")

    def nadji(*kandidati):
        for k in kandidati:
            if k.upper() in zaglavlje:
                return zaglavlje[k.upper()]
        return None

    idx_firma   = nadji("FIRMA", "NAZIV", "NAZIV FIRME")
    idx_adresa  = nadji("ADRESA", "ULICA")
    idx_grad    = nadji("GRAD", "MESTO", "MJESTO")
    idx_pib     = nadji("PIB")
    idx_maticni = nadji("MATIČNI", "MATICNI", "MATIČNI BROJ", "MATICNI BROJ", "MB")

    if idx_firma is None:
        print("\nGreska: Kolona FIRMA nije pronadjena.")
        sys.exit(1)

    print(f"\n  Mapiranje:")
    print(f"    FIRMA   -> naziv_firme")
    print(f"    ADRESA  -> adresa {'(kolona ' + str(idx_adresa + 1) + ')' if idx_adresa is not None else '(nije nadjena)'}")
    print(f"    GRAD    -> {'spaja se sa adresom' if idx_grad is not None else '(nije nadjena)'}")
    print(f"    PIB     -> pib {'(kolona ' + str(idx_pib + 1) + ')' if idx_pib is not None else '(nije nadjena)'}")
    print(f"    MATICNI -> maticni_broj {'(kolona ' + str(idx_maticni + 1) + ')' if idx_maticni is not None else '(nije nadjena)'}")
    print(f"\n  Strategija: azurira postojece po nazivu firme, dodaje nove ako ne postoje.")

    ukupno = ws.max_row - 1
    print(f"\n  Redova za obradu: ~{ukupno}")
    potvrda = unos("Nastavljam? (da/ne)")
    if potvrda.lower() not in ("da", "d", "yes", "y"):
        print("Otkazano.")
        sys.exit(0)

    def cel(row, idx):
        if idx is None:
            return None
        val = row[idx].value
        if isinstance(val, str):
            val = val.strip() or None
        return val

    azurirano = 0
    dodato = 0
    preskoceno = 0
    greske = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        firma = cel(row, idx_firma)
        if not firma:
            preskoceno += 1
            continue

        adresa_val = cel(row, idx_adresa)
        grad_val   = cel(row, idx_grad)
        pib_val    = cel(row, idx_pib)
        maticni_val = cel(row, idx_maticni)

        # Spoji adresu i grad
        if adresa_val and grad_val:
            adresa_puna = f"{adresa_val}, {grad_val}"
        elif adresa_val:
            adresa_puna = adresa_val
        elif grad_val:
            adresa_puna = grad_val
        else:
            adresa_puna = None

        try:
            postojeci = conn.execute(
                "SELECT id FROM klijenti WHERE naziv_firme = ? AND arhiviran = 0",
                (firma,)
            ).fetchone()

            if postojeci:
                conn.execute("""
                    UPDATE klijenti
                    SET adresa = COALESCE(?, adresa),
                        pib = COALESCE(?, pib),
                        maticni_broj = COALESCE(?, maticni_broj)
                    WHERE id = ?
                """, (adresa_puna, pib_val, maticni_val, postojeci[0]))
                azurirano += 1
            else:
                conn.execute("""
                    INSERT INTO klijenti (tip, naziv_firme, adresa, pib, maticni_broj)
                    VALUES ('Pravno', ?, ?, ?, ?)
                """, (firma, adresa_puna, pib_val, maticni_val))
                dodato += 1

        except sqlite3.Error as e:
            print(f"  Red {row_idx}: Greska — {e}")
            greske += 1

    conn.commit()
    conn.close()

    print(f"\n{'=' * 55}")
    print(f"  Gotovo!")
    print(f"  Azurirano klijenata : {azurirano}")
    print(f"  Novih klijenata     : {dodato}")
    print(f"  Praznih redova      : {preskoceno}")
    print(f"  Gresaka             : {greske}")
    print(f"{'=' * 55}")
    input("\nPritisni Enter za izlaz...")


if __name__ == "__main__":
    main()
