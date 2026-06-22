"""
AutoServis — Uvoz vozila i vlasnika iz Excel-a
Pokretanje: python uvoz_vozila.py
Potrebno: pip install openpyxl

Ocekivani format Excel-a (zaglavlje u prvom redu):
  RB | VOZILO | ŠASIJA | TABLICE | (prazno/ID) | VLASNIK | UREĐAJ

Skripta za svaki red:
  1. Kreira klijenta ako jos ne postoji (po nazivu/imenu)
  2. Kreira vozilo i odmah ga vezuje za tog klijenta
"""

import sqlite3
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Nedostaje biblioteka. Pokreni: pip install openpyxl")
    sys.exit(1)


APPDATA = os.environ.get("APPDATA", os.path.expanduser("~"))
DEFAULT_DB = os.path.join(APPDATA, "AutoServis", "autoservis.db")

# Kljucne reci koje oznacavaju pravno lice
PRAVNO_LICE_RECI = {
    "DOO", "D.O.O", "D.O.O.", "AD", "A.D.", "A.D",
    "SUR", "STR", "ZTR", "OU", "OD", "PREDUZEĆE",
    "PREDUZECE", "FIRMA", "AGENCIJA", "KOOPERATIVA",
    "ZADRUGA", "FONDACIJA", "UDRUŽENJE", "UDRUZENJE",
}


def detektuj_tip(naziv):
    """Vrati 'Pravno' ili 'Fizicko' na osnovu naziva."""
    if not naziv:
        return "Fizicko"
    gornji = naziv.upper()
    for rec in PRAVNO_LICE_RECI:
        if rec in gornji:
            return "Pravno"
    return "Fizicko"


def split_ime_prezime(naziv):
    """'MARKO PETROVIC' -> ime='MARKO', prezime='PETROVIC'"""
    delovi = naziv.strip().split()
    if len(delovi) == 1:
        return delovi[0], None
    return delovi[0], " ".join(delovi[1:])


def split_marka_model(vrednost, reci_marke):
    """'SKODA SCALA' -> ('SKODA', 'SCALA') za reci_marke=1"""
    if not vrednost:
        return None, None
    delovi = str(vrednost).strip().split()
    if len(delovi) <= reci_marke:
        return " ".join(delovi), None
    return " ".join(delovi[:reci_marke]), " ".join(delovi[reci_marke:])


def nadji_ili_kreiraj_klijenta(conn, naziv):
    """Vraca klijent_id. Kreira novog klijenta ako ne postoji."""
    naziv = naziv.strip() if naziv else None
    if not naziv:
        return None

    tip = detektuj_tip(naziv)

    if tip == "Pravno":
        row = conn.execute(
            "SELECT id FROM klijenti WHERE naziv_firme = ? AND arhiviran = 0",
            (naziv,)
        ).fetchone()
        if row:
            return row[0]
        conn.execute(
            "INSERT INTO klijenti (tip, naziv_firme) VALUES (?, ?)",
            ("Pravno", naziv)
        )
        return conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    else:
        ime, prezime = split_ime_prezime(naziv)
        row = conn.execute(
            "SELECT id FROM klijenti WHERE ime = ? AND (prezime = ? OR prezime IS NULL) AND arhiviran = 0",
            (ime, prezime)
        ).fetchone()
        if row:
            return row[0]
        conn.execute(
            "INSERT INTO klijenti (tip, ime, prezime) VALUES (?, ?, ?)",
            ("Fizicko", ime, prezime)
        )
        return conn.execute("SELECT last_insert_rowid()").fetchone()[0]


def unos(tekst, podrazumevano=None):
    if podrazumevano:
        odg = input(f"{tekst} [{podrazumevano}]: ").strip()
        return odg if odg else podrazumevano
    return input(f"{tekst}: ").strip()


def main():
    print("=" * 60)
    print("  AutoServis — Uvoz vozila i vlasnika")
    print("=" * 60)

    # --- Baza ---
    print(f"\nPodrazumevana baza: {DEFAULT_DB}")
    db_path = unos("Pritisni Enter za podrazumevano, ili unesi drugu putanju", DEFAULT_DB)
    if not os.path.exists(db_path):
        print(f"\nGreska: Baza nije pronadjena: {db_path}")
        sys.exit(1)

    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON")

    # Dodaj napomena kolonu ako ne postoji
    try:
        conn.execute("ALTER TABLE vozila ADD COLUMN napomena TEXT")
        conn.commit()
        print("  Dodata kolona 'napomena' u tabelu vozila.")
    except sqlite3.OperationalError:
        pass  # Vec postoji

    print("  OK — baza ucitana.")

    # --- Excel ---
    print()
    excel_path = unos("Putanja do Excel fajla").strip('"').strip("'")
    if not os.path.exists(excel_path):
        print(f"\nGreska: Fajl nije pronadjen: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path)
    sheets = wb.sheetnames

    if len(sheets) == 1:
        sheet_name = sheets[0]
        print(f"  Koristim sheet: {sheet_name}")
    else:
        print(f"\nSheet-ovi: {', '.join(sheets)}")
        sheet_name = unos("Koji sheet?")

    ws = wb[sheet_name]

    # Procitaj zaglavlje i pronadji indekse kolona
    zaglavlje = {str(c.value).strip().upper(): i for i, c in enumerate(ws[1]) if c.value}

    print(f"\n  Pronadjene kolone: {', '.join(zaglavlje.keys())}")

    # Mapiraj kolone (fleksibilno — prihvata razlicite nazive)
    def nadji_kolonu(*kandidati):
        for k in kandidati:
            if k.upper() in zaglavlje:
                return zaglavlje[k.upper()]
        return None

    idx_vozilo   = nadji_kolonu("VOZILO", "VOZILA", "AUTO", "AUTOMOBIL")
    idx_sasija   = nadji_kolonu("ŠASIJA", "SASIJA", "ŠASIJA", "VIN", "BROJ ŠASIJE", "BROJ SASIJE")
    idx_tablice  = nadji_kolonu("TABLICE", "REGISTRACIJA", "REG", "TABLICA")
    idx_vlasnik  = nadji_kolonu("VLASNIK", "KLIJENT", "FIRMA", "VLASNICI")
    idx_uredjaj  = nadji_kolonu("UREĐAJ", "UREDJAJ", "NAPOMENA", "OPIS")

    print(f"\n  Mapiranje kolona:")
    print(f"    VOZILO   -> kolona {idx_vozilo + 1 if idx_vozilo is not None else 'NIJE NADJENA'}")
    print(f"    ŠASIJA   -> kolona {idx_sasija + 1 if idx_sasija is not None else 'NIJE NADJENA'}")
    print(f"    TABLICE  -> kolona {idx_tablice + 1 if idx_tablice is not None else 'NIJE NADJENA'}")
    print(f"    VLASNIK  -> kolona {idx_vlasnik + 1 if idx_vlasnik is not None else 'NIJE NADJENA'}")
    print(f"    UREĐAJ   -> kolona {idx_uredjaj + 1 if idx_uredjaj is not None else 'nije mapirana'}")

    if idx_vozilo is None or idx_vlasnik is None:
        print("\nGreska: Kolone VOZILO i VLASNIK su obavezne.")
        sys.exit(1)

    # --- Split marka+model ---
    print()
    while True:
        reci_str = unos("Koliko reci je marka vozila? (1 = SKODA/VW/BMW, 2 = ALFA ROMEO)", "1")
        if reci_str in ("1", "2"):
            reci_marke = int(reci_str)
            break
        print("  Unesi 1 ili 2.")

    # --- Potvrda ---
    ukupno = ws.max_row - 1
    print(f"\n  Redova za uvoz: ~{ukupno}")
    potvrda = unos("Nastavljam? (da/ne)")
    if potvrda.lower() not in ("da", "d", "yes", "y"):
        print("Otkazano.")
        sys.exit(0)

    # --- Uvoz ---
    def cel(row, idx):
        if idx is None:
            return None
        val = row[idx].value
        if isinstance(val, str):
            val = val.strip() or None
        return val

    ubaceno_klijenti = 0
    nadjen_klijenti = 0
    ubaceno_vozila = 0
    greske = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        vlasnik = cel(row, idx_vlasnik)
        vozilo_raw = cel(row, idx_vozilo)

        if not vlasnik and not vozilo_raw:
            continue  # prazan red

        # Klijent
        try:
            postojao = bool(vlasnik and conn.execute(
                "SELECT id FROM klijenti WHERE naziv_firme=? OR (ime=? AND (prezime=? OR prezime IS NULL))",
                (vlasnik, vlasnik.split()[0] if vlasnik else "", " ".join(vlasnik.split()[1:]) if vlasnik and len(vlasnik.split()) > 1 else "")
            ).fetchone())

            klijent_id = nadji_ili_kreiraj_klijenta(conn, vlasnik)

            if postojao:
                nadjen_klijenti += 1
            else:
                ubaceno_klijenti += 1
        except Exception as e:
            print(f"  Red {row_idx}: Greska pri klijentu — {e}")
            greske += 1
            continue

        # Vozilo
        marka, model = split_marka_model(vozilo_raw, reci_marke)
        sasija   = cel(row, idx_sasija)
        tablice  = cel(row, idx_tablice)
        uredjaj  = cel(row, idx_uredjaj)

        try:
            conn.execute(
                """INSERT INTO vozila (klijent_id, marka, model, registracija, broj_sasije, napomena)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (klijent_id, marka, model, tablice, sasija, uredjaj)
            )
            ubaceno_vozila += 1
        except sqlite3.Error as e:
            print(f"  Red {row_idx}: Greska pri vozilu — {e}")
            greske += 1

    conn.commit()
    conn.close()

    print(f"\n{'=' * 60}")
    print(f"  Gotovo!")
    print(f"  Novih klijenata  : {ubaceno_klijenti}")
    print(f"  Klijenata nadjen : {nadjen_klijenti}  (vozilo dodato postoj. klijentu)")
    print(f"  Ubaceno vozila   : {ubaceno_vozila}")
    print(f"  Gresaka          : {greske}")
    print(f"{'=' * 60}")
    input("\nPritisni Enter za izlaz...")


if __name__ == "__main__":
    main()
